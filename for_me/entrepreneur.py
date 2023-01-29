from aiohttp import ClientSession
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_workbook
from fake_useragent import UserAgent

import asyncio


class Graber:
    def __init__(self, url: str):
        self.url: str = url
        self.row_for_one: int = 2
        self.row_for_table: int = 2
        self.headers: dict ={
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'user-agent': UserAgent().random
        }

    async def get_response(self, url: str = None) -> str:
        async with ClientSession(headers=self.headers) as session:
            async with session.get(url if url else self.url) as response:
                return await response.text(encoding='UTF-8')
            
    async def get_tables(self, paginate: int = None) -> str:
        if paginate:
            url = self.url + f'/{paginate}'
            response = await self.get_response(url=url)
        else:
            response = await self.get_response()
        html_response = BeautifulSoup(response, 'lxml')
        tables = html_response.find_all('table', attrs={'class': 'w-full bg-white shadow overflow-hidden sm:rounded-md table-fixed'})[0:5]

        info = []
        for table in tables:
            tasks = []
            tbody = table.find('tbody')

            cards_normal = self.get_trs_normal(tbody=tbody)
            cards_gray = self.get_trs_gray(tbody=tbody)

            cards = cards_normal + cards_gray
            for card in cards:
                number = self.get_number(card=card)
                name = self.get_name(card=card)
                description = self.get_description(card=card)
                investment = self.get_investment(card=card)
                url = self.get_url(card=card)

                info.append([number, name, description, investment, url])
                tasks.append(self.get_one_data(url=url))

            asyncio.gather(*tasks)
        return info


    async def get_one_data(self, url: str):
        response = await self.get_response(url=url)

        html_response = BeautifulSoup(response, 'lxml')

        table = html_response.find('div', attrs={'class': 'bg-white shadow pb-2 sm:rounded-lg mb-12'})
        table = table.find('div', attrs={'class': 'border-t border-gray-200'})

        dd = table.find_all('dd', attrs={'class': 'mt-1 text-base text-gray-900 font-normal sm:mt-0 sm:col-span-2'})
        dt = table.find_all('dt', attrs={'class': 'text-base font-light text-gray-600'})

        attribs = []
        names = []

        for name in dd:
            names.append(name.text.strip())

        for attrib in dt:
            attribs.append(attrib.text.strip())

        data = dict(zip(attribs, names))

        self.write_to_excel(data=data)

    async def get_paginate(self) -> int:
        response = await self.get_response()

        html_response = BeautifulSoup(response, 'lxml')

        paginate = html_response.find('div', attrs={'class': 'overflow-hidden sm:rounded-b-md bg-white overflow-hidden px-4 py-3 flex items-center justify-between sm:px-6'})
        paginate = paginate.find('nav', attrs={'class': 'relative z-0 inline-flex shadow-sm -space-x-px'}).find_all('a')[-2].text.strip()

        return int(paginate)

    def get_trs_normal(self, tbody: BeautifulSoup):
        return tbody.find_all('tr', attrs={'class': 'border-b border-grey-100 hover:bg-blue-50 cursor-pointer'})
    
    def get_trs_gray(self, tbody: BeautifulSoup):
        return tbody.find_all('tr', attrs={'class': 'border-b border-grey-100 hover:bg-blue-50 cursor-pointer bg-gray-50'})
    
    def remove_unwanted_tag(self, *tags: tuple[str], root: BeautifulSoup) -> None:
        if len(tags) > 1:
            for i in tags:
                root.find(i).extract()
        else:
            root.find(tags[0]).extract()

    def get_number(self, card: BeautifulSoup):
        if card.find_all('p', attrs={'class': 'text-lg font-medium text-gray-700 mr-2 w-16'}):
            return card.find('p', attrs={'class': 'text-lg font-medium text-gray-700 mr-2 w-16'}).text
        return 'no numbering'
    
    def get_name(self, card: BeautifulSoup):
        name = card.find('p', attrs={'class': 'text-base font-medium text-gray-700 w-1/2'})

        if name.find_all('br'):
            self.remove_unwanted_tag('br', 'span', root=name)

        return name.text.strip()
    
    def get_description(self, card: BeautifulSoup):
        return card.find('td', attrs={'class': 'hidden lg:table-cell'}).\
            find('p', attrs={'class': 'text-sm font-medium text-gray-700'}).text.strip()

    def get_investment(self, card: BeautifulSoup):
        return card.find('td', attrs={'class': 'hidden md:table-cell w-28'}).\
            find('p', attrs={'class': 'text-sm text-gray-700'}).text.strip()

    def get_url(self, card: BeautifulSoup):
        return 'https://www.entrepreneur.com' + card.\
            find('a', attrs={'class': 'flex items-center w-full'}).get('href')
    
    def write_to_excel(self, data: list or dict):
        if isinstance(data, list):
            try:
                wb = load_workbook('test.xlsx')
                sheet = wb.get_sheet_by_name(name='TableInfo')
            except Exception as ex:
                print(ex)

                wb = Workbook()
                sheet = wb.create_sheet('TableInfo')

            sheet['A1'] = 'Number'
            sheet['B1'] = 'Name'
            sheet['C1'] = 'Description'
            sheet['D1'] = 'Investment'
            sheet['E1'] = 'Url'

            for info in data:
                for card in info:
                    sheet.cell(row=self.row_for_table, column=1, value=card[0])
                    sheet.cell(row=self.row_for_table, column=2, value=card[1])
                    sheet.cell(row=self.row_for_table, column=3, value=card[2])
                    sheet.cell(row=self.row_for_table, column=4, value=card[3])
                    sheet.cell(row=self.row_for_table, column=5, value=card[4])
                    self.row_for_table += 1

        if isinstance(data, dict):
            try:
                wb = load_workbook('test.xlsx')
                sheet =wb.get_sheet_by_name(name='TableInfo')
            except Exception as ex:
                print(ex)

                wb = Workbook()
                sheet = wb.create_sheet('TableInfo')

            sheet['M1'] = 'Industry'
            sheet['N1'] = 'Category'
            sheet['O1'] = 'Founded'
            sheet['P1'] = 'Parent company'
            sheet['Q1'] = 'Leadership'
            sheet['R1'] = 'Address'

            sheet.cell(row=self.row_for_one, column=13, value=data.get('Industry', 'None'))
            sheet.cell(row=self.row_for_one, column=14, value=data.get('Related Categories', 'None'))
            sheet.cell(row=self.row_for_one, column=15, value=data.get('Founded', 'None'))
            sheet.cell(row=self.row_for_one, column=16, value=data.get('Parent Company', 'None'))
            sheet.cell(row=self.row_for_one, column=17, value=data.get('Leadership', 'None'))
            sheet.cell(row=self.row_for_one, column=18, value=data.get('Corporate Address', 'None'))

            self.row_for_one += 1

        save_workbook(wb, 'test.xlsx')

async def main():
    g = Graber(url='https://www.entrepreneur.com/franchises/category/personal-care-businesses')
    paginate = await g.get_paginate()
    data = []
    for i in range(1, paginate+1):
        info = await g.get_tables(i)
        data.append(info)

    g.write_to_excel(data=data)

if __name__ == '__main__':
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())
